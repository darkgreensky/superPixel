import os

import cv2
import numpy as np
import openpyxl

import libs.ERS.demoERS
import libs.SIN.run_demo
from tqdm import tqdm

outer_folder = "C:\\Users\Administrator\Desktop\BSDS"

seg_folders = [
]

img_path = ""
seg_path = []
save_path = ""

importance_info = []
algorithm = ["slic", "seeds", "lsc", "sin", "ers"]
evalution = ["co", "ue", "br", "bp", "asa"]
res = [[[] for __ in evalution] for _ in algorithm]

def init(seg_folder):
    global img_path, save_path
    seg_path.clear()
    img_name = ""
    for item in os.listdir(seg_folder):
        if item[-3:] == "seg":
            # print(seg_folder + item)
            seg_path.append(seg_folder + item)
        if item[-3:] == "jpg":
            img_path = seg_folder + item
            img_name = item[:-4]

    save_path = seg_folder + "res" + "_" + img_name + ".xlsx"


def get_num_pixel(label):
    unique_labels, counts = np.unique(label, return_counts=True)
    num_superpixels = len(unique_labels)
    return num_superpixels


def main():
    folders = os.listdir(outer_folder)
    for item in folders:
        seg_folders.append(outer_folder + "\\" + item + "\\")
    for i in tqdm(range(len(seg_folders)), desc="进度"):
        solve(seg_folders[i])
    print("importance_info:", importance_info)
    print(res)
    save_to_sheet()

def save_to_sheet():
    workbook = openpyxl.Workbook()
    print("save")
    for i in range(len(algorithm)):
        print(algorithm[i])
        sheet = workbook.create_sheet(title=algorithm[i])
        sheet.cell(1, 1, "紧凑度")
        for j in range(len(res[i][0])):
            sheet.cell(1, j + 3, res[i][0][j])
        # sheet.cell(1, 7, "=AVERAGE(B1:F1)")
        sheet.cell(2, 1, "欠分割误差")
        for j in range(len(res[i][1])):
            sheet.cell(2, j + 3, res[i][1][j])
        # sheet.cell(2, 7, "=AVERAGE(B2:F2)")
        sheet.cell(3, 1, "边界召回率")
        for j in range(len(res[i][2])):
            sheet.cell(3, j + 3, res[i][2][j])
        # sheet.cell(3, 7, "=AVERAGE(B3:F3)")
        sheet.cell(4, 1, "边界精度")
        for j in range(len(res[i][3])):
            sheet.cell(4, j + 3, res[i][3][j])
        # sheet.cell(4, 7, "=AVERAGE(B4:F4)")
        sheet.cell(5, 1, "可达分割精度")
        for j in range(len(res[i][4])):
            sheet.cell(5, j + 3, res[i][4][j])
        # sheet.cell(5, 7, "=AVERAGE(B5:F5)")
    if 'Sheet' in workbook.sheetnames:
        std_sheet = workbook['Sheet']
        workbook.remove(std_sheet)
    workbook.save(outer_folder + "\\" + "res.xlsx")

def solve(seg_folder):
    # img_path, _ = QFileDialog.getOpenFileName(None, "Open File", "", "All Files (*);;Text Files (*.jpg)")
    init(seg_folder)
    # print(img_path)
    img = cv2.imdecode(np.fromfile(img_path, dtype=np.uint8), -1)
    # print("img read")
    # print("slic")

    slic = cv2.ximgproc.createSuperpixelSLIC(img, algorithm=cv2.ximgproc.SLIC, region_size=15)
    slic.iterate(10)
    slic_label = slic.getLabels()

    # print("slico")
    # slico = cv2.ximgproc.createSuperpixelSLIC(img, algorithm=cv2.ximgproc.SLICO, region_size=15)
    # slico.iterate(10)
    # slico_label = slico.getLabels()
    #
    # print("mslic")
    # mslic_list = []
    # region_size_mslic = 10
    # while True:
    #     print(region_size_mslic, end=' ')
    #     mslic = cv2.ximgproc.createSuperpixelSLIC(img, algorithm=cv2.ximgproc.MSLIC, region_size=region_size_mslic)
    #     mslic.iterate(10)
    #     mslic_label = mslic.getLabels()
    #     num = get_num_pixel(mslic_label)
    #     print(num)
    #     mslic_list.append((abs(num - 650), region_size_mslic))
    #     if num > 3000:
    #         region_size_mslic += 10
    #     elif num > 2000:
    #         region_size_mslic += 5
    #     elif num < 300:
    #         region_size_mslic += 10
    #     elif num < 350:
    #         region_size_mslic += 5
    #     elif num < 400:
    #         region_size_mslic += 3
    #     else:
    #         region_size_mslic += 1
    #     if region_size_mslic >= 150:
    #         mslic_list = sorted(mslic_list)
    #         print(mslic_list[0][0], mslic_list[0][1])
    #         importance_info.append((mslic_list[0][0], mslic_list[0][1]))
    #         mslic = cv2.ximgproc.createSuperpixelSLIC(img, algorithm=cv2.ximgproc.MSLIC, region_size=mslic_list[0][1])
    #         mslic.iterate(10)
    #         mslic_label = mslic.getLabels()
    #         break

    # print("seeds")
    seeds = cv2.ximgproc.createSuperpixelSEEDS(img.shape[1], img.shape[0], img.shape[2], num_superpixels=2000,
                                               num_levels=15, prior=3, histogram_bins=5, double_step=True)
    seeds.iterate(img, 10)
    seeds_label = seeds.getLabels()

    # print("lsc")
    lsc = cv2.ximgproc.createSuperpixelLSC(img, 15, 0.075)
    lsc.iterate(10)
    lsc_label = lsc.getLabels()

    # print("SIN")
    result_img, label_SIN, toc = libs.SIN.run_demo.SIN_handle(img)
    # print("ERS")
    label_ERS, toc = libs.ERS.demoERS.ERS_handle(img, 650)
    label = [slic_label, seeds_label, lsc_label, label_SIN, label_ERS]
    segs = []
    # print("read human label")
    for seg in seg_path:
        true_labels = read_human_segments_label_file(seg)
        segs.append(true_labels)
    workbook = openpyxl.Workbook()
    for i in range(len(algorithm)):
        # print("----------------------------------")
        print(algorithm[i])
        # print("----------------------------------")
        # sheet = workbook.create_sheet(title=algorithm[i])
        # sheet = workbook.active
        # sheet.cell(1, 1, "紧凑度")
        # sheet.cell(1, 7, "=AVERAGE(B1:F1)")
        # sheet.cell(2, 1, "欠分割误差")
        # sheet.cell(2, 7, "=AVERAGE(B2:F2)")
        # sheet.cell(3, 1, "边界召回率")
        # sheet.cell(3, 7, "=AVERAGE(B3:F3)")
        # sheet.cell(4, 1, "边界精度")
        # sheet.cell(4, 7, "=AVERAGE(B4:F4)")
        # sheet.cell(5, 1, "可达分割精度")
        # sheet.cell(5, 7, "=AVERAGE(B5:F5)")
        # print("co")
        for j in range(len(seg_path)):
            co = calculate_compactness(label[i])
            # print("co:", co)
            res[i][0].append(co)
            # sheet.cell(1, j + 2, co)

        # print("ue")
        for j in range(len(seg_path)):
            ue = compute_undersegmentation_error(label[i], segs[j])
            # print("ue:", ue)
            res[i][1].append(ue)
            # sheet.cell(2, j + 2, ue)

        # print("br")
        for j in range(len(seg_path)):
            br = compute_boundary_recall(label[i], segs[j])
            # print("br:", br)
            res[i][2].append(br)
            # sheet.cell(3, j + 2, br)

        # print("bp")
        for j in range(len(seg_path)):
            bp = compute_boundary_precision(label[i], segs[j])
            # print("bp:", bp)
            res[i][3].append(bp)
            # sheet.cell(4, j + 2, bp)

        # print("asa")
        for j in range(len(seg_path)):
            asa = compute_achievable_segmentation_accuracy(label[i], segs[j])
            # print("asa:", asa)
            res[i][4].append(asa)
            # sheet.cell(5, j + 2, asa)
    # if 'Sheet' in workbook.sheetnames:
    #     std_sheet = workbook['Sheet']
    #     workbook.remove(std_sheet)
    # workbook.save(save_path)


def read_human_segments_label_file(file_name):
    with open(file_name, "r") as f:
        lines = f.readlines()
    width = 0
    height = 0
    num_segments = 0
    i = 0
    while True:
        if lines[i].split()[0] == 'width':
            width = int(lines[i].split()[1])
        elif lines[i].split()[0] == 'height':
            height = int(lines[i].split()[1])
        elif lines[i].split()[0] == 'segments':
            num_segments = int(lines[i].split()[1])
        elif lines[i].split()[0] == 'data':
            i += 1
            break
        i += 1

    segmented_img = np.zeros((height, width, 3), dtype=np.uint8)
    true_labels = np.zeros((height, width), dtype=np.uint16)

    # 设置区域颜色
    colors = np.random.randint(0, 255, size=(num_segments, 3), dtype=np.uint8)

    # 填充图像
    for line in lines[11:]:
        s, y, x1, x2 = map(int, line.split())
        segmented_img[y, x1: x2] = colors[s]
        true_labels[y, x1: x2] = s
    return true_labels


def calculate_compactness(labels):
    superpixels = np.max(labels) + 1

    perimeter = np.zeros(superpixels)
    area = np.zeros(superpixels)

    for i in range(labels.shape[0]):
        for j in range(labels.shape[1]):
            count = 0

            if i > 0:
                if labels[i, j] != labels[i - 1, j]:
                    count += 1
            else:
                count += 1

            if i < labels.shape[0] - 1:
                if labels[i, j] != labels[i + 1, j]:
                    count += 1
            else:
                count += 1

            if j > 0:
                if labels[i, j] != labels[i, j - 1]:
                    count += 1
            else:
                count += 1

            if j < labels.shape[1] - 1:
                if labels[i, j] != labels[i, j + 1]:
                    count += 1
            else:
                count += 1

            perimeter[labels[i, j]] += count
            area[labels[i, j]] += 1

    compactness = 0

    for i in range(superpixels):
        if perimeter[i] > 0:
            compactness += area[i] * (4 * np.pi * area[i]) / (perimeter[i] * perimeter[i])

    compactness /= labels.size

    if compactness > 1.0:
        print("Invalid compactness:", compactness)

    return compactness


def compute_intersection_matrix(labels, gt):
    superpixels = np.max(labels) + 1
    gt_segments = np.max(gt) + 1

    superpixel_sizes = np.zeros(superpixels, dtype=int)
    gt_sizes = np.zeros(gt_segments, dtype=int)

    intersection_matrix = np.zeros((gt_segments, superpixels), dtype=int)

    for i in range(labels.shape[0]):
        for j in range(labels.shape[1]):
            intersection_matrix[gt[i, j], labels[i, j]] += 1
            superpixel_sizes[labels[i, j]] += 1
            gt_sizes[gt[i, j]] += 1

    return intersection_matrix, superpixel_sizes, gt_sizes


def compute_undersegmentation_error(labels, gt):
    H, W = gt.shape
    N = H * W

    intersection_matrix, superpixel_sizes, gt_sizes = compute_intersection_matrix(labels, gt)

    if intersection_matrix is None:
        return None

    error = 0
    for j in range(intersection_matrix.shape[1]):
        min_diff = np.inf
        for i in range(intersection_matrix.shape[0]):
            superpixel_j_minus_gt_i = superpixel_sizes[j] - intersection_matrix[i, j]
            if superpixel_j_minus_gt_i < 0:
                print("Set difference is negative.")
            if superpixel_j_minus_gt_i < min_diff:
                min_diff = superpixel_j_minus_gt_i

        error += min_diff
    return error / N


def compute_boundary_recall(labels, gt, d=0.0025):
    H, W = gt.shape
    r = int(round(d * np.sqrt(H * H + W * W)))
    tp = 0
    fn = 0

    for i in range(H):
        for j in range(W):
            if is_4_connected_boundary_pixel(gt, i, j):
                pos = False
                for k in range(max(0, i - r), min(H - 1, i + r) + 1):
                    for l in range(max(0, j - r), min(W - 1, j + r) + 1):
                        if is_4_connected_boundary_pixel(labels, k, l):
                            pos = True
                if pos:
                    tp += 1
                else:
                    fn += 1

    if tp + fn > 0:
        return tp / (tp + fn)
    return 0


def compute_boundary_precision(labels, gt, d=0.0025):
    H, W = gt.shape
    r = round(d * np.sqrt(H * H + W * W))

    tp = 0
    fp = 0

    for i in range(H):
        for j in range(W):
            if is_4_connected_boundary_pixel(gt, i, j):

                pos = False
                # Search for boundary pixel in the supervoxel segmentation.
                for k in range(max(0, i - r), min(H - 1, i + r) + 1):
                    for l in range(max(0, j - r), min(W - 1, j + r) + 1):
                        if is_4_connected_boundary_pixel(labels, k, l):
                            pos = True

                if pos:
                    tp += 1
            elif is_4_connected_boundary_pixel(labels, i, j):
                pos = False
                # Search for boundary pixel in the supervoxel segmentation.
                for k in range(max(0, i - r), min(H - 1, i + r) + 1):
                    for l in range(max(0, j - r), min(W - 1, j + r) + 1):
                        if is_4_connected_boundary_pixel(gt, k, l):
                            pos = True

                if not pos:
                    fp += 1

    if tp + fp > 0:
        return tp / (tp + fp)

    return 0


def is_4_connected_boundary_pixel(labels, i, j):
    if i > 0:
        if labels[i, j] != labels[i - 1, j]:
            return True

    if i < labels.shape[0] - 1:
        if labels[i, j] != labels[i + 1, j]:
            return True

    if j > 0:
        if labels[i, j] != labels[i, j - 1]:
            return True

    if j < labels.shape[1] - 1:
        if labels[i, j] != labels[i, j + 1]:
            return True

    return False


def compute_achievable_segmentation_accuracy(labels, gt):
    H, W = gt.shape[:2]
    N = H * W

    intersection_matrix, superpixel_sizes, gt_sizes = compute_intersection_matrix(labels, gt)

    accuracy = 0
    for j in range(intersection_matrix.shape[1]):
        max_intersection = np.max(intersection_matrix[:, j])
        accuracy += max_intersection

    return accuracy / N


if __name__ == '__main__':
    main()
